import openpyxl
from openpyxl.styles import Alignment

# Class that is responsable for creating excel files, at a specific location, and writting inside them the wanted data
class ExcelCreator:

    def set_column_width_to_fit_header(self, sheet):
        for col_idx, cell in enumerate(sheet[1], start=1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = len(str(cell.value))
            adjusted_width = max_length + 6
            sheet.column_dimensions[column_letter].width = adjusted_width

    def create_excel(self, header_row, data_rows, file_path):
        # Creating a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # The top row is the header row, so we BOLD each entry
        sheet.append(header_row)
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

        # Add all the entries to the excel
        for row in data_rows:
            sheet.append(row)
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

        # Set column width to fit the header text
        self.set_column_width_to_fit_header(sheet)

        # Freeze the header row using A1 notation
        sheet.freeze_panes = 'A2'

        # Save the workbook to the specified file path
        workbook.save(file_path)
        print("Excel file created successfully at:", file_path)
